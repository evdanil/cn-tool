from __future__ import annotations

import json
from dataclasses import fields, is_dataclass
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING, Any, Mapping

import yaml
from openpyxl import Workbook

from cn_lens.models import (
    InvalidLensObject,
    LensFinding,
    LensObject,
    LensResult,
    LensRun,
    ObjectSet,
)

if TYPE_CHECKING:
    from cn_lens.reports.aggregate import LensReport


# Workflow names whose summary blocks receive per-workflow rendering treatment.
# When a LensResult's summary dict has one of these keys the renderer emits a
# fenced code block (markdown) or a Per-Workflow row (xlsx).
_WORKFLOW_SUMMARY_KEYS: frozenset[str] = frozenset(
    {
        "impact",
        "dns",
        "reachability",
        "device",
        "validate_site",
        "decommission_site",
        "allocate",
        "config_find",
        "config_diff",
        "report",
        "bssid",
        "stats",
        "e911",
    }
)


def _plain_value(value: Any) -> Any:
    if isinstance(value, Enum):
        return value.value
    if is_dataclass(value) and not isinstance(value, type):
        return {
            field.name: _plain_value(getattr(value, field.name))
            for field in fields(value)
        }
    if isinstance(value, Mapping):
        return {str(key): _plain_value(item) for key, item in value.items()}
    if isinstance(value, tuple):
        return [_plain_value(item) for item in value]
    if isinstance(value, list):
        return [_plain_value(item) for item in value]
    return value


def _object_to_dict(lens_object: LensObject) -> dict[str, Any]:
    return {
        "original": lens_object.original,
        "normalized": lens_object.normalized,
        "object_type": lens_object.object_type.value,
        "value": lens_object.value,
        "notes": list(lens_object.notes),
    }


def _invalid_to_dict(invalid: InvalidLensObject) -> dict[str, str]:
    return {
        "original": invalid.original,
        "reason": invalid.reason,
    }


def _object_set_to_dict(inputs: ObjectSet) -> dict[str, Any]:
    return {
        "objects": [_object_to_dict(lens_object) for lens_object in inputs.objects],
        "invalid": [_invalid_to_dict(invalid) for invalid in inputs.invalid],
        "duplicate_count": inputs.duplicate_count,
    }


def _finding_to_dict(finding: LensFinding) -> dict[str, Any]:
    return {
        "severity": finding.severity,
        "source": finding.source,
        "message": finding.message,
        "detail": _plain_value(finding.detail),
    }


def _result_to_dict(result: LensResult) -> dict[str, Any]:
    return {
        "lens_object": _object_to_dict(result.lens_object),
        "status": result.status,
        "summary": _plain_value(result.summary),
        "findings": [_finding_to_dict(finding) for finding in result.findings],
        "sources": _plain_value(result.sources),
    }


def run_to_dict(run: LensRun) -> dict[str, Any]:
    return {
        "schema_version": run.schema_version,
        "tool": run.tool,
        "workflow": run.workflow,
        "run_id": run.run_id,
        "inputs": _object_set_to_dict(run.inputs),
        "results": [_result_to_dict(result) for result in run.results],
        "warnings": list(run.warnings),
        "errors": list(run.errors),
    }


def render_json(run: LensRun) -> str:
    return json.dumps(run_to_dict(run), indent=2, sort_keys=False) + "\n"


def render_yaml(run: LensRun) -> str:
    return yaml.safe_dump(run_to_dict(run), sort_keys=False)


def render_markdown(run: LensRun) -> str:
    lines = [
        f"# {run.tool} {run.workflow} report",
        "",
        "## Summary",
        "",
        f"- Schema version: {run.schema_version}",
        f"- Run ID: {run.run_id}",
        f"- Objects: {len(run.inputs.objects)}",
        f"- Invalid inputs: {len(run.inputs.invalid)}",
        f"- Duplicate inputs: {run.inputs.duplicate_count}",
        f"- Results: {len(run.results)}",
    ]
    if run.warnings:
        lines.extend(["", "## Warnings", ""])
        lines.extend(f"- {warning}" for warning in run.warnings)
    if run.errors:
        lines.extend(["", "## Errors", ""])
        lines.extend(f"- {error}" for error in run.errors)
    if run.inputs.invalid:
        lines.extend(["", "## Invalid Inputs", ""])
        lines.extend(
            f"- `{invalid.original}`: {invalid.reason}" for invalid in run.inputs.invalid
        )
    if run.results:
        lines.extend(["", "## Results", ""])
        for result in run.results:
            obj = result.lens_object
            lines.extend(
                [
                    f"### {obj.value}",
                    "",
                    f"- Type: {obj.object_type.value}",
                    f"- Status: {result.status}",
                ]
            )
            if result.sources:
                lines.append("- Sources: " + _join_mapping(result.sources))
            if result.summary:
                workflow_blocks = _extract_workflow_summary_blocks(result.summary)
                plain_summary = _plain_value(result.summary)
                if workflow_blocks:
                    # Emit non-workflow summary keys inline (if any remain)
                    non_workflow = {
                        k: v
                        for k, v in plain_summary.items()  # type: ignore[union-attr]
                        if k not in _WORKFLOW_SUMMARY_KEYS
                    }
                    if non_workflow:
                        lines.append(
                            "- Summary: "
                            + json.dumps(non_workflow, sort_keys=True)
                        )
                    # Emit each workflow-specific block as a fenced YAML block
                    for wf_key, wf_value in _plain_value(workflow_blocks).items():
                        lines.extend(
                            [
                                "",
                                f"#### Workflow: {wf_key}",
                                "",
                                "```yaml",
                                yaml.safe_dump(
                                    {wf_key: wf_value}, sort_keys=False
                                ).rstrip(),
                                "```",
                            ]
                        )
                else:
                    lines.append(
                        "- Summary: "
                        + json.dumps(plain_summary, sort_keys=True)
                    )
            if result.findings:
                lines.append("- Findings:")
                lines.extend(
                    f"  - {finding.severity} {finding.source}: {finding.message}"
                    for finding in result.findings
                )
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _join_mapping(values: Mapping[str, Any]) -> str:
    return ", ".join(f"{key}: {values[key]}" for key in sorted(values))


def _extract_workflow_summary_blocks(
    summary: Mapping[str, Any],
) -> dict[str, Any]:
    """Return the subset of *summary* whose keys are known workflow names."""
    return {
        key: value
        for key, value in summary.items()
        if key in _WORKFLOW_SUMMARY_KEYS
    }


def _xlsx_safe_value(value: Any) -> Any:
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _xlsx_safe_row(values: tuple[Any, ...]) -> tuple[Any, ...]:
    return tuple(_xlsx_safe_value(value) for value in values)


def _append_xlsx_row(sheet: Any, values: tuple[Any, ...]) -> None:
    sheet.append(_xlsx_safe_row(values))


def _format_kv_block(items: Mapping[str, Any], indent: str = "    ") -> list[str]:
    """Return indented `key: value` lines with values column-aligned.

    Key followed by colon (no space before colon, matches `key: value`
    convention). Values are right-padded by post-colon spaces so they line
    up visually. Nested mappings/lists are pretty-printed via YAML instead
    of dumped as JSON.
    """
    if not items:
        return []
    keys = list(items.keys())
    width = max(len(str(k)) for k in keys)
    lines: list[str] = []
    for key in keys:
        value = items[key]
        pad = " " * (width - len(str(key)))
        if isinstance(value, Mapping) or (
            isinstance(value, (list, tuple)) and not isinstance(value, str)
        ):
            rendered = yaml.safe_dump(_plain_value(value), sort_keys=False).rstrip()
            lines.append(f"{indent}{key}:")
            for sub in rendered.splitlines():
                lines.append(f"{indent}  {sub}")
        else:
            lines.append(f"{indent}{key}:{pad} {value}")
    return lines


_SEVERITY_TAGS: Mapping[str, str] = {
    "error": "ERR ",
    "warn": "WARN",
    "warning": "WARN",
    "info": "INFO",
    "debug": "DBG ",
}


def render_text(run: LensRun) -> str:
    """Plain-text report — ASCII only, no ANSI. Grep-friendly.

    Layout: framed header, key-value summary block, sectioned bullets for
    warnings/errors/invalid inputs, per-result block with sources, summary,
    and findings broken out as readable key-value pairs (no JSON dumps).
    """
    bar = "=" * 72
    sub_bar = "-" * 72
    lines: list[str] = [
        bar,
        f"  {run.tool} {run.workflow} report",
        bar,
    ]
    header = {
        "run_id": run.run_id,
        "schema_version": run.schema_version,
        "objects": len(run.inputs.objects),
        "invalid_inputs": len(run.inputs.invalid),
        "duplicate_inputs": run.inputs.duplicate_count,
        "results": len(run.results),
        "warnings": len(run.warnings),
        "errors": len(run.errors),
    }
    lines.extend(_format_kv_block(header, indent=""))

    if run.warnings:
        lines.extend(["", "warnings:"])
        lines.extend(f"  - {warning}" for warning in run.warnings)
    if run.errors:
        lines.extend(["", "errors:"])
        lines.extend(f"  - {error}" for error in run.errors)
    if run.inputs.invalid:
        lines.extend(["", "invalid inputs:"])
        for invalid in run.inputs.invalid:
            lines.append(f"  - {invalid.original}")
            lines.append(f"    reason: {invalid.reason}")

    if run.results:
        lines.extend(["", "results:"])
        for result in run.results:
            obj = result.lens_object
            lines.append("")
            lines.append(f"  {sub_bar}")
            lines.append(
                f"  {obj.value}  [{obj.object_type.value}]  status: {result.status}"
            )
            lines.append(f"  {sub_bar}")
            if obj.original and obj.original != obj.value:
                lines.append(f"    original    : {obj.original}")
            if obj.normalized and obj.normalized != obj.value:
                lines.append(f"    normalized  : {obj.normalized}")
            if obj.notes:
                lines.append(f"    notes       : {'; '.join(obj.notes)}")
            if result.sources:
                lines.append("    sources:")
                src_keys = sorted(result.sources)
                src_w = max(len(k) for k in src_keys)
                for k in src_keys:
                    pad = " " * (src_w - len(k))
                    lines.append(f"      {k}:{pad} {result.sources[k]}")
            if result.summary:
                lines.append("    summary:")
                lines.extend(
                    _format_kv_block(_plain_value(result.summary), indent="      ")
                )
            if result.findings:
                lines.append("    findings:")
                for finding in result.findings:
                    tag = _SEVERITY_TAGS.get(finding.severity.lower(), finding.severity.upper())
                    lines.append(
                        f"      [{tag}] {finding.source}: {finding.message}"
                    )
                    if finding.detail:
                        for sub in _format_kv_block(
                            _plain_value(finding.detail), indent="          "
                        ):
                            lines.append(sub)
    return "\n".join(lines) + "\n"


_SEVERITY_STYLES: Mapping[str, str] = {
    "error": "bold red",
    "warn": "yellow",
    "warning": "yellow",
    "info": "cyan",
    "debug": "dim",
}

_STATUS_STYLES: Mapping[str, str] = {
    "classified": "cyan",
    "ok": "green",
    "found": "green",
    "matched": "green",
    "not_found": "yellow",
    "missing": "yellow",
    "warn": "yellow",
    "partial": "yellow",
    "error": "bold red",
    "failed": "bold red",
    "unsupported": "red",
    "invalid": "red",
    "not_queried": "dim",
    "not_configured": "dim",
    "disabled": "dim",
}


def render_human(run: LensRun, *, force_ansi: bool = True) -> str:
    """Rich-formatted report with ANSI color, tables, and panels.

    Designed for reading on a terminal. For grep/automation use ``--format
    text`` instead. When writing to a file (``--output``), pass
    ``force_ansi=False`` to suppress ANSI escape codes so the file contains
    plain text suitable for downstream tools (diff, grep, email).
    """
    from io import StringIO
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.text import Text

    buf = StringIO()
    if force_ansi:
        console = Console(file=buf, force_terminal=True, color_system="truecolor", width=100)
    else:
        console = Console(file=buf, force_terminal=False, no_color=True, width=100)

    title = Text(f"{run.tool} {run.workflow} report", style="bold white on blue")
    console.print(Panel.fit(title, border_style="blue"))

    header_table = Table.grid(padding=(0, 2))
    header_table.add_column(style="bold cyan", justify="right")
    header_table.add_column()
    header_table.add_row("run_id", run.run_id)
    header_table.add_row("schema", str(run.schema_version))
    header_table.add_row("objects", str(len(run.inputs.objects)))
    header_table.add_row("invalid", str(len(run.inputs.invalid)))
    header_table.add_row("duplicates", str(run.inputs.duplicate_count))
    header_table.add_row("results", str(len(run.results)))
    header_table.add_row(
        "warnings",
        Text(str(len(run.warnings)), style="yellow" if run.warnings else "dim"),
    )
    header_table.add_row(
        "errors",
        Text(str(len(run.errors)), style="bold red" if run.errors else "dim"),
    )
    console.print(header_table)

    if run.warnings:
        console.print()
        console.print(Text("warnings", style="bold yellow"))
        for warning in run.warnings:
            console.print(Text(f"  • {warning}", style="yellow"))
    if run.errors:
        console.print()
        console.print(Text("errors", style="bold red"))
        for error in run.errors:
            console.print(Text(f"  • {error}", style="red"))
    if run.inputs.invalid:
        console.print()
        console.print(Text("invalid inputs", style="bold red"))
        for invalid in run.inputs.invalid:
            console.print(
                Text("  • ", style="red")
                + Text(invalid.original, style="bold")
                + Text(f"  ({invalid.reason})", style="dim")
            )

    if run.results:
        console.print()
        for result in run.results:
            obj = result.lens_object
            status_style = _STATUS_STYLES.get(result.status.lower(), "white")
            head = (
                Text(obj.value, style="bold")
                + Text(f"  [{obj.object_type.value}]  ", style="dim")
                + Text(f"status: {result.status}", style=status_style)
            )
            body_lines: list[Any] = []

            if obj.original and obj.original != obj.value:
                body_lines.append(Text(f"original   : {obj.original}", style="dim"))
            if obj.normalized and obj.normalized != obj.value:
                body_lines.append(Text(f"normalized : {obj.normalized}", style="dim"))
            if obj.notes:
                body_lines.append(Text(f"notes      : {'; '.join(obj.notes)}", style="dim"))

            if result.sources:
                src_table = Table(
                    show_header=False, box=None, padding=(0, 1), expand=False
                )
                src_table.add_column(style="cyan", justify="right")
                src_table.add_column()
                for key in sorted(result.sources):
                    val = str(result.sources[key])
                    style = _STATUS_STYLES.get(val.lower(), "white")
                    src_table.add_row(key, Text(val, style=style))
                body_lines.append(Text("sources:", style="bold"))
                body_lines.append(src_table)

            if result.summary:
                summary_table = Table(
                    show_header=False, box=None, padding=(0, 1), expand=False
                )
                summary_table.add_column(style="cyan", justify="right")
                summary_table.add_column()
                plain_summary = _plain_value(result.summary)
                for key, value in plain_summary.items():
                    if isinstance(value, (Mapping, list, tuple)) and not isinstance(value, str):
                        rendered = yaml.safe_dump(value, sort_keys=False).rstrip()
                        summary_table.add_row(str(key), Text(rendered, style="white"))
                    else:
                        summary_table.add_row(str(key), Text(str(value), style="white"))
                body_lines.append(Text("summary:", style="bold"))
                body_lines.append(summary_table)

            if result.findings:
                body_lines.append(Text("findings:", style="bold"))
                for finding in result.findings:
                    style = _SEVERITY_STYLES.get(finding.severity.lower(), "white")
                    body_lines.append(
                        Text(f"  [{finding.severity.upper()}] ", style=style)
                        + Text(f"{finding.source}: ", style="cyan")
                        + Text(finding.message, style="white")
                    )

            from rich.console import Group
            border = _STATUS_STYLES.get(result.status.lower(), "white")
            console.print(Panel(Group(*body_lines), title=head, title_align="left", border_style=border))

    return buf.getvalue()


_SUBNET_DETAIL_BASE_COLUMNS: tuple[str, ...] = (
    "Original Input",
    "IP",
    "Mask",
    "Name",
    "MAC",
    "DHCP",
    "DHCP Scope Start",
    "DHCP Scope End",
    "DHCP Servers",
    "DHCP Options\nOption - Value",
    "DHCP Options\nOption - Decoded Value",
    "DHCP Failover Association",
    "Notes",
    "Inherited Fields",
)


def _write_subnet_data_detail(workbook: Any, run: LensRun) -> None:
    """Append a 'Subnet Data Detail' sheet when any result has deep infoblox data.

    Mirrors the layout of the ``modules/subnet_request.py:_save_subnet_data``
    "Subnet Data Detail" sheet.  Only created when at least one result contains
    a deep infoblox block (``summary["infoblox"]["deep"] == True``).

    Row structure (one prefix = one main row + secondary rows per sub-object):
    - Main row: subnet identity + DHCP scope + member/option summary + Notes
    - DNS record rows: one per in-subnet DNS record (IP + /32 Mask + A Record)
    - Fixed address rows: one per fixedaddress (IP + /32 Mask + Name + MAC)
    """
    detail_rows: list[tuple[Any, ...]] = []

    for result in run.results:
        ib = result.summary.get("infoblox", {})
        if not ib.get("deep"):
            continue

        obj_value = result.lens_object.value
        network = ib.get("network", obj_value)

        # Split network into IP + mask components (e.g. "10.0.0.0/24" → "10.0.0.0", "/24")
        if "/" in network:
            ip_part, mask_suffix = network.split("/", 1)
            mask_part = f"/{mask_suffix}"
        else:
            ip_part, mask_part = network, ""

        dhcp_ranges = ib.get("dhcp_ranges", [])
        members = ib.get("members", [])
        dhcp_options = ib.get("dhcp_options", [])
        dns_records = ib.get("dns_records", [])
        fixed_addresses = ib.get("fixed_addresses", [])

        first_range = dhcp_ranges[0] if dhcp_ranges else {}
        is_dhcp = "Y" if (members or dhcp_ranges) else "N"

        # DHCP servers summary: "name - IP" per member
        dhcp_servers_str = "\n".join(
            f"{m.get('name', '')} - {m.get('ip', '')}" for m in members
        )

        # DHCP options value summary
        dhcp_options_value_str = "\n".join(
            f"{o.get('name', '')} - {o.get('value', '')}" for o in dhcp_options
        )
        # DHCP options decoded value summary
        has_decoded = any(o.get("decoded_value") for o in dhcp_options)
        dhcp_options_decoded_str = (
            "\n".join(
                f"{o.get('name', '')} - {o.get('decoded_value', '')}"
                for o in dhcp_options
            )
            if has_decoded
            else ""
        )

        failover_str = first_range.get("failover_association", "") if first_range else ""
        comment = ib.get("comment", "")

        main_row = (
            obj_value,          # Original Input
            ip_part,            # IP
            mask_part,          # Mask
            "Subnet",           # Name
            "",                 # MAC
            is_dhcp,            # DHCP
            first_range.get("start_addr", "") if first_range else "",  # DHCP Scope Start
            first_range.get("end_addr", "") if first_range else "",    # DHCP Scope End
            dhcp_servers_str,   # DHCP Servers
            dhcp_options_value_str,    # DHCP Options\nOption - Value
            dhcp_options_decoded_str,  # DHCP Options\nOption - Decoded Value
            failover_str,       # DHCP Failover Association
            comment,            # Notes
            "",                 # Inherited Fields
        )
        detail_rows.append(main_row)

        # DNS record rows
        for rec in dns_records:
            rec_ip = rec.get("ip", "")
            detail_rows.append((
                "",             # Original Input (blank for secondary rows)
                rec_ip,         # IP
                "/32",          # Mask
                rec.get("name", ""),  # Name (A Record)
                "",             # MAC
                "",             # DHCP
                "", "",         # Scope Start / End
                "",             # DHCP Servers
                "", "",         # Options
                "",             # Failover
                "DNS record",   # Notes
                "",             # Inherited Fields
            ))

        # Fixed address rows
        for fa in fixed_addresses:
            detail_rows.append((
                "",             # Original Input
                fa.get("ip", ""),   # IP
                "/32",          # Mask
                fa.get("name", ""),  # Name
                fa.get("mac", ""),   # MAC
                "",             # DHCP
                "", "",         # Scope Start / End
                "",             # DHCP Servers
                "", "",         # Options
                "",             # Failover
                "Fixed IP",     # Notes
                "",             # Inherited Fields
            ))

    if not detail_rows:
        return

    detail_sheet = workbook.create_sheet("Subnet Data Detail")
    _append_xlsx_row(detail_sheet, _SUBNET_DETAIL_BASE_COLUMNS)
    for row in detail_rows:
        _append_xlsx_row(detail_sheet, row)


def write_xlsx(run: LensRun, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _append_xlsx_row(summary, ("Field", "Value"))
    _append_xlsx_row(summary, ("schema_version", run.schema_version))
    _append_xlsx_row(summary, ("tool", run.tool))
    _append_xlsx_row(summary, ("workflow", run.workflow))
    _append_xlsx_row(summary, ("run_id", run.run_id))
    _append_xlsx_row(summary, ("objects", len(run.inputs.objects)))
    _append_xlsx_row(summary, ("invalid_inputs", len(run.inputs.invalid)))
    _append_xlsx_row(summary, ("duplicate_inputs", run.inputs.duplicate_count))
    _append_xlsx_row(summary, ("warnings", len(run.warnings)))
    _append_xlsx_row(summary, ("errors", len(run.errors)))

    inputs = workbook.create_sheet("Inputs")
    _append_xlsx_row(
        inputs,
        ("kind", "original", "normalized", "object_type", "value", "reason", "notes"),
    )
    for obj in run.inputs.objects:
        _append_xlsx_row(
            inputs,
            (
                "object",
                obj.original,
                obj.normalized,
                obj.object_type.value,
                obj.value,
                "",
                "; ".join(obj.notes),
            ),
        )
    for invalid in run.inputs.invalid:
        _append_xlsx_row(
            inputs,
            ("invalid", invalid.original, "", "invalid", "", invalid.reason, ""),
        )

    results = workbook.create_sheet("Results")
    _append_xlsx_row(results, ("object", "object_type", "status", "summary", "sources"))
    for result in run.results:
        _append_xlsx_row(
            results,
            (
                result.lens_object.value,
                result.lens_object.object_type.value,
                result.status,
                json.dumps(_plain_value(result.summary), sort_keys=True),
                json.dumps(_plain_value(result.sources), sort_keys=True),
            ),
        )

    findings = workbook.create_sheet("Findings")
    _append_xlsx_row(findings, ("object", "severity", "source", "message", "detail"))
    for result in run.results:
        for finding in result.findings:
            _append_xlsx_row(
                findings,
                (
                    result.lens_object.value,
                    finding.severity,
                    finding.source,
                    finding.message,
                    json.dumps(_plain_value(finding.detail), sort_keys=True),
                ),
            )

    # --- Per-Workflow sheet (only when workflow-specific summary keys exist) ---
    per_workflow_rows: list[tuple[str, str, str, str]] = []
    for result in run.results:
        workflow_blocks = _extract_workflow_summary_blocks(result.summary)
        if not workflow_blocks:
            continue
        plain_blocks = _plain_value(workflow_blocks)
        for wf_key, wf_value in plain_blocks.items():
            if isinstance(wf_value, dict):
                for k, v in wf_value.items():
                    str_v = (
                        json.dumps(v, sort_keys=True)
                        if isinstance(v, (dict, list))
                        else str(v)
                    )
                    per_workflow_rows.append(
                        (result.lens_object.value, wf_key, str(k), str_v)
                    )
            elif isinstance(wf_value, list):
                for item in wf_value:
                    if isinstance(item, dict):
                        for k, v in item.items():
                            str_v = (
                                json.dumps(v, sort_keys=True)
                                if isinstance(v, (dict, list))
                                else str(v)
                            )
                            per_workflow_rows.append(
                                (result.lens_object.value, wf_key, str(k), str_v)
                            )
                    else:
                        per_workflow_rows.append(
                            (result.lens_object.value, wf_key, "", str(item))
                        )
            else:
                per_workflow_rows.append(
                    (result.lens_object.value, wf_key, "", str(wf_value))
                )

    if per_workflow_rows:
        pw_sheet = workbook.create_sheet("Per-Workflow")
        _append_xlsx_row(pw_sheet, ("object", "workflow", "key", "value"))
        for row in per_workflow_rows:
            _append_xlsx_row(pw_sheet, row)

    # --- Subnet Data Detail sheet (parity with modules/subnet_request.py) ---
    _write_subnet_data_detail(workbook, run)

    workbook.save(path)


def render_report(report: "LensReport", fmt: str, output: Path | None = None) -> str | None:
    """Render a ``LensReport`` aggregate in the requested format.

    Parameters
    ----------
    report:
        The ``LensReport`` to render.
    fmt:
        Output format string.  Same set as ``render_run`` plus aliases.
    output:
        When provided, write to file and return ``None``.  When ``None``,
        return the rendered string.

    Returns
    -------
    str | None
        Rendered text, or ``None`` when *output* is supplied.

    Raises
    ------
    ValueError
        Unknown format string, or xlsx requested without *output*.
    OSError
        Write failure when *output* is supplied.
    """
    # Lazy import to break the renderers ↔ aggregate circular dependency.
    # Imported once here and passed as an argument to the inner helpers that
    # require it, avoiding three separate identical local imports.
    from cn_lens.reports.aggregate import report_to_dict

    normalized_fmt = fmt.lower()
    if normalized_fmt == "yml":
        normalized_fmt = "yaml"

    if normalized_fmt == "xlsx":
        if output is None:
            raise ValueError("xlsx output requires an output path")
        _write_report_xlsx(report, output)
        return None

    report_dict = report_to_dict(report)

    renderers_map = {
        "json": _render_report_json,
        "yaml": _render_report_yaml,
        "markdown": _render_report_markdown,
        "md": _render_report_markdown,
        "text": _render_report_text,
        "txt": _render_report_text,
        "human": _render_report_text,
    }
    renderer = renderers_map.get(normalized_fmt)
    if renderer is None:
        raise ValueError(f"unknown render format: {fmt}")

    rendered = renderer(report, report_dict)
    if output is None:
        return rendered
    output.write_text(rendered, encoding="utf-8")
    return None


def _render_report_json(report: "LensReport", report_dict: dict[str, Any]) -> str:
    return json.dumps(report_dict, indent=2, sort_keys=False) + "\n"


def _render_report_yaml(report: "LensReport", report_dict: dict[str, Any]) -> str:
    return yaml.safe_dump(report_dict, sort_keys=False)


def _render_report_markdown(report: "LensReport", report_dict: dict[str, Any]) -> str:
    lines = [
        "# cn-lens report",
        "",
        "## Summary",
        "",
        f"- Report ID: {report.report_id}",
        f"- Tool: {report.tool}",
        f"- Schema version: {report.schema_version}",
        f"- Generated at: {report.generated_at}",
        f"- Runs included: {len(report.runs)}",
    ]
    if report.findings:
        lines.extend(["", "## Findings", ""])
        for finding in report.findings:
            lines.append(
                f"- {finding.severity} {finding.source}: {finding.message}"
            )
    for run in report.runs:
        lines.extend([
            "",
            f"## Run {run.run_id} ({run.workflow})",
            "",
            f"- Tool: {run.tool}",
            f"- Workflow: {run.workflow}",
            f"- Objects: {len(run.inputs.objects)}",
            f"- Results: {len(run.results)}",
        ])
        if run.warnings:
            lines.append(f"- Warnings: {len(run.warnings)}")
        if run.errors:
            lines.append(f"- Errors: {len(run.errors)}")
    return "\n".join(lines).rstrip() + "\n"


def _render_report_text(report: "LensReport", report_dict: dict[str, Any]) -> str:
    lines = [
        "cn-lens report",
        f"report_id: {report.report_id}",
        f"tool: {report.tool}",
        f"schema_version: {report.schema_version}",
        f"generated_at: {report.generated_at}",
        f"runs: {len(report.runs)}",
    ]
    if report.findings:
        lines.extend(["", "findings:"])
        for finding in report.findings:
            lines.append(
                f"- {finding.severity} {finding.source}: {finding.message}"
            )
    for run in report.runs:
        lines.extend([
            "",
            f"run: {run.run_id} ({run.workflow})",
            f"  tool: {run.tool}",
            f"  objects: {len(run.inputs.objects)}",
            f"  results: {len(run.results)}",
        ])
    return "\n".join(lines) + "\n"


def _write_report_xlsx(report: "LensReport", path: Path) -> None:
    """Write a LensReport to an xlsx workbook with per-run sheets."""
    workbook = Workbook()

    # --- Summary sheet ---
    summary_sheet = workbook.active
    summary_sheet.title = "Report Summary"
    _append_xlsx_row(summary_sheet, ("Field", "Value"))
    _append_xlsx_row(summary_sheet, ("schema_version", report.schema_version))
    _append_xlsx_row(summary_sheet, ("tool", report.tool))
    _append_xlsx_row(summary_sheet, ("report_id", report.report_id))
    _append_xlsx_row(summary_sheet, ("generated_at", report.generated_at))
    _append_xlsx_row(summary_sheet, ("runs_count", len(report.runs)))
    _append_xlsx_row(summary_sheet, ("findings_count", len(report.findings)))

    # --- Per-run sheets ---
    for run in report.runs:
        # Create a short sheet title slug from the run_id (max 31 chars for xlsx)
        slug = run.run_id[-15:] if len(run.run_id) > 15 else run.run_id
        sheet_title = f"{slug}"[:31]

        run_sheet = workbook.create_sheet(title=sheet_title)
        _append_xlsx_row(run_sheet, ("Field", "Value"))
        _append_xlsx_row(run_sheet, ("run_id", run.run_id))
        _append_xlsx_row(run_sheet, ("workflow", run.workflow))
        _append_xlsx_row(run_sheet, ("tool", run.tool))
        _append_xlsx_row(run_sheet, ("objects", len(run.inputs.objects)))
        _append_xlsx_row(run_sheet, ("results", len(run.results)))
        _append_xlsx_row(run_sheet, ("warnings", len(run.warnings)))
        _append_xlsx_row(run_sheet, ("errors", len(run.errors)))

        # Results sub-table
        run_sheet.append([])
        _append_xlsx_row(run_sheet, ("object", "object_type", "status", "summary"))
        for result in run.results:
            _append_xlsx_row(
                run_sheet,
                (
                    result.lens_object.value,
                    result.lens_object.object_type.value,
                    result.status,
                    json.dumps(_plain_value(result.summary), sort_keys=True),
                ),
            )

    workbook.save(path)


# ---------------------------------------------------------------------------
# P5.2 — config find xlsx writer (per-device tabs + =HYPERLINK formulas)
# ---------------------------------------------------------------------------

#: Maximum number of per-device config tabs to write (mirrors cn-tool cap).
_CONFIG_FIND_MAX_DEVICE_TABS = 50


def write_config_find_xlsx(
    result: Any,
    path: Path,
    *,
    runtime: Any = None,
) -> None:
    """Write a ``MultiTermResult`` to an xlsx workbook.

    Sheet layout
    ------------
    ``Matches``
        One row per ``ConfigMatch``.  The ``Line #`` cell contains an
        ``=HYPERLINK("#'DEVICE'!A<row>", <line_no>)`` formula that jumps to
        the matching line in the device's config sheet.
    ``<DEVICE>`` (up to ``_CONFIG_FIND_MAX_DEVICE_TABS`` device tabs)
        Full config content of each matched device file, one line per row.
        Only created when the source file is readable and under the
        ``report_max_config_tab_kb`` size limit (default 512 KB).

    Parameters
    ----------
    result:
        A ``MultiTermResult`` from ``adapters.config_repo.search_multi_term``.
    path:
        Destination ``*.xlsx`` file path.
    runtime:
        Optional runtime used to read ``cfg["report_max_config_tab_kb"]``.
    """
    from openpyxl import Workbook

    workbook = Workbook()

    # ---- Matches sheet -----
    matches_sheet = workbook.active
    matches_sheet.title = "Matches"
    _append_xlsx_row(
        matches_sheet,
        ("Search Term", "Device", "Line #", "Line Content", "File Path"),
    )

    # Collect unique devices with their file paths (preserve insertion order)
    device_file_map: dict[str, str] = {}
    for match in result.matches:
        if match.device not in device_file_map:
            device_file_map[match.device] = match.file_path

    # F6 (donor parity): when >50 matched devices, write only the Matches sheet
    # with plain line numbers — no per-device config tabs and no HYPERLINK formulas.
    # This mirrors modules/config_search.py:473-475 which returns after the Matches
    # sheet when device count exceeds the cap (all-or-nothing behaviour).
    over_cap = len(device_file_map) > _CONFIG_FIND_MAX_DEVICE_TABS

    # Build a term index for grouping — needed for the Matches sheet regardless of cap.
    import re as _re

    term_for_match: list[str] = []
    terms_list = list(result.term_results.keys()) if result.term_results else []
    compiled_terms: list[Any] = []
    for t in terms_list:
        try:
            compiled_terms.append((t, _re.compile(t, _re.IGNORECASE)))
        except _re.error:
            compiled_terms.append((t, None))

    for match in result.matches:
        assigned = ""
        for term, pat in compiled_terms:
            if pat is not None and pat.search(match.snippet):
                assigned = term
                break
            elif pat is None and term.lower() in match.snippet.lower():
                assigned = term
                break
        term_for_match.append(assigned)

    if over_cap:
        # Plain Matches sheet only — no HYPERLINK formulas, no per-device tabs.
        for i, match in enumerate(result.matches):
            term_label = term_for_match[i] if i < len(term_for_match) else ""
            _append_xlsx_row(
                matches_sheet,
                (term_label, match.device, match.line_number, match.snippet, match.file_path),
            )
        workbook.save(path)
        return

    # Tabs + hyperlinks path (<=50 devices).
    # Determine which devices will get tabs (cap at 50)
    tabbed_devices: set[str] = set(list(device_file_map.keys())[:_CONFIG_FIND_MAX_DEVICE_TABS])

    # Map device name → (1-based row number in the device tab) for each match
    # so we can build HYPERLINK formulas.  We track line offsets per device sheet.
    # Each device sheet row 1 = header; row 2+ = config lines.
    # HYPERLINK formula: =HYPERLINK("#'DEVICE'!A<row_in_device_sheet>", line_no)
    # Since device sheets have one config line per row (starting at row 1, no
    # header), the sheet row = match.line_number + 1 (1-based).

    for i, match in enumerate(result.matches):
        term_label = term_for_match[i] if i < len(term_for_match) else ""
        if match.device in tabbed_devices:
            # Sheet row = line_number + 1 (0-based line → 1-based row, no header in device sheet)
            sheet_row = match.line_number + 1
            # hyperlink_formula is the only raw formula cell; all other cells are
            # sanitised via _xlsx_safe_value to prevent formula injection (F4).
            hyperlink_formula = f'=HYPERLINK("#\'{match.device}\'!A{sheet_row}", {match.line_number})'
            matches_sheet.append(
                [
                    _xlsx_safe_value(term_label),
                    _xlsx_safe_value(match.device),
                    hyperlink_formula,  # intentional formula — keep raw
                    _xlsx_safe_value(match.snippet),
                    _xlsx_safe_value(match.file_path),
                ]
            )
        else:
            _append_xlsx_row(
                matches_sheet,
                (term_label, match.device, match.line_number, match.snippet, match.file_path),
            )

    # ---- Per-device config sheets ----
    cfg = {}
    if runtime is not None and hasattr(runtime, "cfg"):
        cfg = runtime.cfg or {}
    max_config_tab_kb = int(cfg.get("report_max_config_tab_kb", 512) or 512)
    max_config_tab_bytes = max(1, max_config_tab_kb) * 1024

    for device in list(device_file_map.keys())[:_CONFIG_FIND_MAX_DEVICE_TABS]:
        file_path = Path(device_file_map[device])
        try:
            file_size = file_path.stat().st_size
        except OSError:
            # File not accessible — write a notice tab
            dev_sheet = workbook.create_sheet(title=device[:31])
            _append_xlsx_row(dev_sheet, (f"Config file not found: {device_file_map[device]}",))
            continue

        if file_size > max_config_tab_bytes:
            dev_sheet = workbook.create_sheet(title=device[:31])
            _append_xlsx_row(
                dev_sheet,
                (
                    f"Config omitted: size {round(file_size / 1024, 1)} KB exceeds "
                    f"limit {max_config_tab_kb} KB.",
                ),
            )
            continue

        try:
            lines = file_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        except OSError:
            dev_sheet = workbook.create_sheet(title=device[:31])
            _append_xlsx_row(dev_sheet, (f"Error reading config: {device_file_map[device]}",))
            continue

        dev_sheet = workbook.create_sheet(title=device[:31])
        for line in lines:
            # Protect formula injection
            safe = _xlsx_safe_value(line)
            dev_sheet.append([safe])

    workbook.save(path)


def _write_config_find_run_xlsx(run: LensRun, path: Path, *, runtime: Any = None) -> None:
    """Dispatch a config_find LensRun to the specialised xlsx writer.

    Reconstructs a ``MultiTermResult``-compatible object from the per-result
    ``summary["config_find"]["cfg_matches"]`` dicts so that
    :func:`write_config_find_xlsx` can build the Matches sheet and per-device
    config tabs.

    Only cfg_matches (config-repo hits) are included in the specialised xlsx;
    yaml_matches are omitted because they reference YAML paths, not line-numbered
    config files, and the writer's hyperlink formula assumes line-numbered tabs.
    """
    from cn_lens.adapters.config_repo import ConfigMatch, MultiTermResult

    all_matches: list[ConfigMatch] = []
    for result in run.results:
        cf = result.summary.get("config_find", {})
        for m_dict in cf.get("cfg_matches", []):
            try:
                all_matches.append(ConfigMatch(
                    device=m_dict.get("device", ""),
                    file_path=m_dict.get("file_path", ""),
                    line_number=int(m_dict.get("line_number", 0)),
                    snippet=m_dict.get("snippet", ""),
                    context_before=tuple(m_dict.get("context_before", ())),
                    context_after=tuple(m_dict.get("context_after", ())),
                ))
            except (TypeError, ValueError):
                # Malformed entry — skip rather than crash.
                pass

    # Use the query terms (object values) as the term keys for the term_results dict.
    # Consume real term_status and source_status stored in the summary (F2 fix).
    term_results: dict[str, Any] = {}
    source_status = "live"
    for result in run.results:
        cf = result.summary.get("config_find", {})
        term = result.lens_object.value
        # Use persisted term_status when available; fall back to computed count.
        stored_term_stat = cf.get("term_status")
        if stored_term_stat is not None:
            term_results[term] = stored_term_stat
        else:
            n_matches = len([
                m for m in all_matches
                if any(
                    m_d.get("device") == m.device and m_d.get("line_number") == m.line_number
                    for m_d in cf.get("cfg_matches", [])
                )
            ])
            term_results[term] = {"matched": n_matches, "missed": n_matches == 0}
        # Take the most specific source_status ("indexed" wins over "live").
        stored_ss = cf.get("source_status", "live")
        if stored_ss == "indexed":
            source_status = "indexed"

    multi_term = MultiTermResult(
        matches=tuple(all_matches),
        total_files_scanned=sum(
            result.summary.get("config_find", {}).get("total_cfg_files_scanned", 0)
            for result in run.results
        ),
        truncated=any(
            result.summary.get("config_find", {}).get("truncated", False)
            for result in run.results
        ),
        term_results=term_results,
        source_status=source_status,
    )

    write_config_find_xlsx(multi_term, path, runtime=runtime)


def render_run(run: LensRun, fmt: str, output: Path | None = None) -> str | None:
    normalized_fmt = fmt.lower()
    if normalized_fmt == "yml":
        normalized_fmt = "yaml"
    if normalized_fmt == "xlsx":
        if output is None:
            raise ValueError("xlsx output requires an output path")
        if run.workflow == "config_find":
            _write_config_find_run_xlsx(run, output)
        else:
            write_xlsx(run, output)
        return None

    renderers = {
        "json": render_json,
        "yaml": render_yaml,
        "markdown": render_markdown,
        "md": render_markdown,
        "text": render_text,
        "txt": render_text,
        "human": render_human,
    }
    renderer = renderers.get(normalized_fmt)
    if renderer is None:
        raise ValueError(f"unknown render format: {fmt}")

    # Suppress ANSI when rendering to a file so the output is plain text.
    if normalized_fmt == "human" and output is not None:
        rendered = render_human(run, force_ansi=False)
    else:
        rendered = renderer(run)
    if output is None:
        return rendered
    output.write_text(rendered, encoding="utf-8")
    return None
